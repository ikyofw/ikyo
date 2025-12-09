import copy
import logging
import math
import os
import pathlib
import re
import string
import traceback
from plistlib import InvalidFileException

import django.core.files.uploadedfile as djangoUploadedfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from core.core.exception import IkValidateException
from core.core.lang import Boolean2
from core.utils.lang_utils import isNotNullBlank, isNullBlank

logger = logging.getLogger('ikyo')


def columnName2Index(colname):
    colname = colname.upper()
    if type(colname) is not str:
        return colname
    col = 0
    power = 1
    for i in range(len(colname) - 1, -1, -1):
        ch = colname[i]
        col += (ord(ch) - ord('A') + 1) * power
        power *= 26
    return col


def columnIndex2Name(i) -> str:
    nm = ''
    ci = i - 1
    index = ci // 26
    if index > 0:
        nm += columnIndex2Name(index)
    nm += string.ascii_uppercase[ci % 26]
    return nm


def copyCellStyle(fromCell, toCell):
    fromCell.fill = copy.copy(toCell.fill)
    fromCell.font = copy.copy(toCell.font)
    fromCell.number_format = copy.copy(toCell.number_format)
    fromCell.border = copy.copy(toCell.border)
    fromCell._style = copy.copy(toCell._style)
    fromCell.alignment = copy.copy(toCell.alignment)
    fromCell.protection = copy.copy(toCell.protection)


DATA_START_COLUMN_INDEX = 3  # column index starts from 1. "3" = "C"

'''
    Screen definition.
    Please reference to View-v14.xlsx
    Usage example:
        file=r'var/ikyo/screen/office.xlsx'                  
        sp = SpreadsheetParser(file)
        print(sp.data)
'''


class SpreadsheetParser:
    TABLE_END_FLAG = 'End of Table'
    '''
        End of Table
    '''
    END_FLAG = 'EOD'
    '''
        END of Component
    '''

    def __init__(self, excelFile, ignoreBlankLines=True, tableNames=None, ignore_output_data=True) -> None:
        self.__firstDataColumnIndex = DATA_START_COLUMN_INDEX - 1  # starts from 0 . the value is (3-1)=2
        # If the line is blank, then igmore it if set to True
        self.__ignoreBlankLines = ignoreBlankLines
        self.__tableNames = tableNames
        # 1. read data from spreadsheet file

        if isinstance(excelFile, djangoUploadedfile.UploadedFile):
            self.__inputs = {}
            self.__file = excelFile.name
        elif os.path.isfile(excelFile):
            self.__inputs = {}
            self.__file = excelFile
        else:
            raise Exception('[' + excelFile + '] file does not exist.')

        with pd.ExcelFile(excelFile) as f:
            sheetData = []
            for sheetName in f.sheet_names:
                d = pd.read_excel(f, sheet_name=sheetName)
                if len(d.columns) > 0 and d.columns[0] is not None \
                        and type(d.columns[0]) == str \
                        and d.columns[0].lower() == 'do not modify this column':
                    d = d.where(d.notnull(), None)
                    d = d.to_numpy()
                    sheetData.append(d)
            # 2 .read user's input data
            for data in sheetData:
                for i in range(len(data)):
                    row = data[i]
                    prmName = row[0]
                    if prmName is not None and len(prmName) > 0:
                        if prmName.upper() == self.END_FLAG:
                            break
                        if ignore_output_data and prmName[0] == '>':  # ignore output parameters. e.g. ">projectName"
                            continue
                        else:
                            v = self.__readPrm(prmName, data, i)
                            self.__inputs[prmName] = v

    def __readPrm(self, name, data, startRowNo):
        isTable = self.__isTable(name, data, startRowNo)
        v = None
        if not isTable:
            row = data[startRowNo]
            v = row[self.__firstDataColumnIndex]
        else:
            v = []
            titles = data[startRowNo]
            if len(titles) > self.__firstDataColumnIndex:  # table's first column is column "C"
                totalTableColumns = 0
                titles = titles[self.__firstDataColumnIndex:]
                for title in titles:
                    if title is not None and (type(title) == str and title != '' or type(title) == float and not math.isnan(title)):
                        totalTableColumns += 1
                    else:
                        break
                titles = titles[0:totalTableColumns]
                # table data starts from next line
                i = startRowNo + 1
                while True:
                    row = data[i]
                    row = row[self.__firstDataColumnIndex:]
                    endTableFlag = row[0]
                    if endTableFlag is not None and type(endTableFlag) == str \
                            and endTableFlag == self.TABLE_END_FLAG:
                        break
                    rowCellData = row[0:totalTableColumns].tolist()
                    if self.__ignoreBlankLines:
                        isAllBlank = True
                        for item in rowCellData:
                            if not isNullBlank(item):
                                isAllBlank = False
                                break
                        if not isAllBlank:
                            v.append(rowCellData)
                    else:
                        v.append(rowCellData)
                    i += 1
        return v

    def __isTable(self, name, sheetData, startRowNo):
        if self.__tableNames is not None and name in self.__tableNames:
            return True
        # check the end "End of Table" flag
        for i in range(startRowNo + 1, len(sheetData)):
            row = sheetData[i]
            prmName, tableFlag = row[0], row[2]
            if not isNullBlank(prmName):
                return False
            if tableFlag == self.TABLE_END_FLAG:
                return True
        return False

    '''
    def cell_background_is_cyan ( workbook, cell ):
        # Returns TRUE if the given cell from the given workbook has a solid cyan RGB (0,255,255) background.
        # Note that the workbook must be opened with formatting_info = True, i.e.
        #     xlrd.open_workbook(xls_filename, formatting_info=True)
        assert type (cell) is xlrd.sheet.Cell
        assert type (workbook) is xlrd.book.Book

        xf_index = cell.xf_index
        xf_style = workbook.xf_list[xf_index]
        xf_background = xf_style.background

        fill_pattern = xf_background.fill_pattern
        pattern_colour_index = xf_background.pattern_colour_index
        background_colour_index = xf_background.background_colour_index

        pattern_colour = workbook.colour_map[pattern_colour_index]
        background_colour = workbook.colour_map[background_colour_index]

        # If the cell has a solid cyan background, then:
        #  - fill_pattern will be 0x01
        #  - pattern_colour will be cyan (0,255,255)
        #  - background_colour is not used with fill pattern 0x01. (undefined value)
        #    So despite the name, for a solid fill, the background colour is not actually the background colour.
        # Refer https://www.openoffice.org/sc/excelfileformat.pdf S. 2.5.12 'Patterns for Cell and Chart Background Area'
        if fill_pattern == 0x01 and pattern_colour == (0,255,255):
            return True
        return False
    '''

    @property
    def file(self):
        return self.__file

    @property
    def data(self):
        return self.__inputs

    def get(self, name):
        return None if name not in self.__inputs.keys() else self.__inputs[name]

    def __str__(self):
        return str(self.__inputs.keys())


class SpreadsheetWriter:
    def __init__(self, parameters, templateFile, outputFile, check_table_column_amount: bool = True) -> None:
        if not isinstance(templateFile, djangoUploadedfile.UploadedFile):
            if templateFile is None or not os.path.isfile(templateFile):
                logger.error('Template file [%s] is not found.', templateFile)
                raise IkValidateException('Template file not found.')
        if isNullBlank(outputFile):
            logger.error('Output file can not be empty.', outputFile)
            raise IkValidateException('Template file not found.')
        if os.path.isfile(outputFile):
            logger.error('Output file [%s] is exists.', outputFile)
            raise IkValidateException('Output file is exists.')
        self.__parameters = {} if parameters is None else parameters
        self.__templateFile = templateFile
        self.__outputFile = outputFile
        self.__check_table_column_amount = check_table_column_amount

    def write(self) -> Boolean2:
        # 2025-11-26: TODO: images missing issue
        wb = None
        try:
            try:
                if self.__templateFile.suffix.lower() == '.xlsm':
                    wb = load_workbook(self.__templateFile, keep_vba=True)
                else:
                    wb = load_workbook(self.__templateFile)
            except InvalidFileException as e:
                raise IkValidateException(f"Template file format error: {self._templateFile}\n{e}")
            if len(self.__parameters) > 0:
                self.__write(wb)
            if os.path.isfile(self.__outputFile):
                logger.error('Output file [%s] is exists.', self.__outputFile)
                raise IkValidateException('Output file is exists.')
            pathlib.Path(pathlib.Path(self.__outputFile).parent).mkdir(parents=True, exist_ok=True)
            wb.save(self.__outputFile)
            return Boolean2(True, 'success')
        except Exception as e:
            traceback.print_exc()
            logger.error('export failed: %s', e)
            return Boolean2(trueFalse=False, data='System error.')
        finally:
            if wb is not None:
                try:
                    wb.close()
                except:
                    logger.error('Close workbook [%s] failed', self.__outputFile)

    def __write(self, wb):
        # 1. get all parameters in the template file
        nameSheetDict = {}
        for stname in wb.sheetnames:
            st = wb[stname]
            if st[1] is None or st[1][0].value is None or str(st[1][0].value).lower() != 'Do not modify this column'.lower():
                continue
            i = 1   # starts from row 2
            while True:
                i += 1
                if i > st.max_row:
                    break
                cell = st[i][0]
                if cell is not None:
                    content = cell.value
                    if content is not None:
                        if str(content).upper() == 'EOD':
                            break
                        elif content.strip() != '':
                            parameter_name = content.strip()
                            if parameter_name not in nameSheetDict.keys():
                                nameSheetDict[parameter_name] = [st]
                            else:
                                nameSheetDict[parameter_name].append(st)
        for name, value in self.__parameters.items():
            if name not in nameSheetDict.keys():
                print('Parameter [%s] is not found in the template file [%s]' % (name, self.__templateFile))
            sheet_objects = nameSheetDict.get(name, None)
            if sheet_objects is None:
                logger.error('Parameter [%s] is not found in the template file [%s]', name, self.__templateFile)
                continue
            for st in sheet_objects:
                data = value
                sheetName = None
                columnIndex = DATA_START_COLUMN_INDEX
                rowIndex = None
                if type(value) == dict:  # Example: 'segmentation': {'data': segmentation, 'sheet': 'segmentation','rowIndex': 17, 'columnIndex': 'E'}
                    sheetName = value.get('sheet', None)
                    if sheetName is not None:
                        st = wb[sheetName]
                        if st is None:
                            raise IkValidateException('Sheet [%s] does not exist. Please check.' % sheetName)
                    data = value.get('data', None)
                    columnIndex = value.get('columnIndex', DATA_START_COLUMN_INDEX)
                    if columnIndex is not None and type(columnIndex) == str:
                        columnIndex = columnName2Index(columnIndex)
                    rowIndex = value.get('rowIndex', None)
                if st is None:
                    raise IkValidateException('Parameter name [%s] does not exist. Please check.' % name)

                if self.__isSingleValue(st, name, data):
                    self.__writeSingleValue(st, name, data, columnIndex=columnIndex, rowIndex=rowIndex)
                else:
                    table_is_last_area = self.__table_is_last_area(st, name)
                    self.__writeTableValue(st, name, data, columnIndex=columnIndex, rowIndex=rowIndex, is_last_area=table_is_last_area)

    def __table_is_last_area(self, st, name) -> bool:
        """ Check Is the name component the last component
        """
        start_row = None
        # find table row
        for cell in st['A']:
            if cell.value == name:
                start_row = cell.row
                break
        values_below = []
        row = start_row + 1
        while row <= st.max_row:
            value = st.cell(row=row, column=cell.column).value
            if value is not None:
                if len(values_below) == 0 and value == SpreadsheetParser.END_FLAG:
                    return True
                values_below.append(value)
            row += 1
        return False if len(values_below) > 0 else True

    def __isSingleValue(self, st, name, value) -> bool:
        return type(value) != list  # None can be single value or table value

    def __getNameRowIndex(self, st, name) -> int:
        i = 1   # starts from row 2
        while True:
            i += 1
            if i > st.max_row:
                break
            cell = st[i][0]
            if cell is not None:
                content = cell.value
                if content is not None and content == name:
                    return i
        return -1

    def __writeSingleValue(self, st, name, value, columnIndex=DATA_START_COLUMN_INDEX, rowIndex=None):
        rowNo = rowIndex if rowIndex is not None else self.__getNameRowIndex(st, name)
        st[columnIndex2Name(columnIndex) + str(rowNo)] = value

    def __is_merged_but_not_top_left(self, sheet, cell):
        return isinstance(cell, MergedCell)

    def __find_next_writable_cell(self, sheet, row_idx, col_idx):
        max_col = sheet.max_column
        for offset in range(1, max_col - col_idx + 1):
            try_col = col_idx + offset
            cell = sheet.cell(row=row_idx, column=try_col)
            if not isinstance(cell, MergedCell):
                return cell
        return None

    def __writeTableValue(self, st, name, values, columnIndex=DATA_START_COLUMN_INDEX, rowIndex=None, is_last_area=False):
        if values is None or len(values) == 0:
            return
        v2 = [(v if (type(v) == list or type(v) == tuple) else [v]) for v in values]

        lastTitleRow = rowIndex - 1 if rowIndex is not None else self.__getNameRowIndex(st, name)

        # get header merged cells
        header_merges = []
        for merged_range in st.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_row == max_row == lastTitleRow:
                header_merges.append((min_col, max_col))
        tableColumnTitles = []
        firstRowCells = []
        merged_cell_num = 0
        for colIndex in range(columnIndex, len(st[lastTitleRow]) + 1):
            col_idx = colIndex + merged_cell_num
            cell = st[columnIndex2Name(col_idx) + str(lastTitleRow)]
            if self.__is_merged_but_not_top_left(st, cell):
                new_cell = self.__find_next_writable_cell(st, lastTitleRow, col_idx)
                if new_cell and isNotNullBlank(new_cell.value):
                    tableColumnTitles.append(new_cell.value)
                    firstRowCells.append(st[columnIndex2Name(new_cell.column) + str(lastTitleRow + 1)])
                    merged_cell_num += new_cell.column - cell.column
                else:
                    break
            else:
                if cell.value is None or cell.value == '':
                    break
                tableColumnTitles.append(cell.value)
                firstRowCells.append(st[columnIndex2Name(col_idx) + str(lastTitleRow + 1)])
        rowIndex = lastTitleRow + 1

        # clean table data if exists
        rowNo = rowIndex
        totalEmptyRows = 0
        while True:
            tableEndFlag = st[columnIndex2Name(columnIndex) + str(rowNo)].value
            if tableEndFlag == SpreadsheetParser.TABLE_END_FLAG:
                break
            for first_row_cell in firstRowCells:
                cell = st.cell(row=rowIndex, column=first_row_cell.column)
                cell.value = None
            rowNo += 1
            if rowNo > st.max_row or rowNo > 1048576:
                break
            totalEmptyRows += 1

        required_rows = len(v2) - totalEmptyRows  # need insert rows number

        # Completely move the content below the table (including styles and merged cells)
        # Noted: must be processed before inserting data into the table
        if required_rows > 0 and not is_last_area:
            self.__shift_tail_area(st, move_start_row=rowIndex + totalEmptyRows, move_count=required_rows)  # Move down all areas after the table by required_rows rows
            self.__shift_print_area(st, row_offset=required_rows)  # Move down the print area by required_rows rows

        for i in range(len(v2)):
            rowData = v2[i]
            if self.__check_table_column_amount and len(rowData) != len(tableColumnTitles):
                logger.error("Table %s Row %s data length [%s] does not match table column length [%s]." % (name, i + 1, len(rowData), len(tableColumnTitles)))
                raise IkValidateException("Table %s Row %s data length does not match table column length." % (name, i + 1))
            if i > 0:   # template table has a blank row
                rowIndex += 1
                if totalEmptyRows == 1:
                    # is not last table, insert at __shift_tail_area method
                    if is_last_area:
                        st.insert_rows(rowIndex)
                    # copy style and cells merge
                    self.__hand_row_style_merges(st, rowIndex - 1, rowIndex, header_merges)
                else:
                    totalEmptyRows -= 1

            for j, first_row_cell in enumerate(firstRowCells):
                if j >= len(rowData):
                    break
                cell_data = rowData[j]
                cell = st.cell(row=rowIndex, column=first_row_cell.column)
                cell.value = cell_data
            if i > 0:    # update the cell styles (starts from the 2nd row)
                for j, first_row_cell in enumerate(firstRowCells):
                    if j >= len(rowData):
                        break
                    cellName = columnIndex2Name(first_row_cell.column) + str(rowIndex)
                    copyCellStyle(st[cellName], first_row_cell)

    def __hand_row_style_merges(self, st, src_row_idx, dst_row_idx, header_merges):
        # copy style
        for col in range(1, st.max_column + 1):
            src_cell = st.cell(row=src_row_idx, column=col)
            dst_cell = st.cell(row=dst_row_idx, column=col)
            copyCellStyle(dst_cell, src_cell)
        # copy merge cells
        for min_col, max_col in header_merges:
            merge_range = f"{get_column_letter(min_col)}{dst_row_idx}:{get_column_letter(max_col)}{dst_row_idx}"
            st.merge_cells(merge_range)

    def __shift_tail_area(self, st: Worksheet, move_start_row: int, move_count: int):
        """ Move down the specified area by the specified row

        Args:
            st (Worksheet): sheet
            move_start_row (int): start row number
            move_count (int): rows offset number
        """
        max_col = st.max_column
        max_row = st.max_row

        # Save area content and style
        temp_rows = []
        for r in range(max_row, move_start_row - 1, -1):  # Handle from bottom to top to avoid coverage
            row_data = []
            for c in range(1, max_col + 1):
                cell = st.cell(row=r, column=c)
                row_data.append((cell.value, copy.copy(cell._style)))
            temp_rows.insert(0, row_data)  # Insert to the front and maintain order

        # Save merged cells
        merge_to_shift = []
        for merge in list(st.merged_cells.ranges):
            if merge.min_row >= move_start_row:
                merge_to_shift.append(merge)
                st.unmerge_cells(str(merge))

        # insert rows
        st.insert_rows(move_start_row, amount=move_count)

        # Restore content and style
        for i, row_data in enumerate(temp_rows):
            for j, (val, style) in enumerate(row_data):
                cell = st.cell(row=move_start_row + move_count + i, column=j + 1)
                cell.value = val
                cell._style = style

        # Restore merged cells
        for merge in merge_to_shift:
            merge.shift(0, move_count)
            if merge is not None:
                try:
                    st.merge_cells(merge.coord)
                except Exception as e:
                    logger.warning(f"Skip invalid merged cells {merge.coord} - {e}")
            else:
                logger.warning(f"Skip empty merged cells.")

    def __shift_print_area(self, st, row_offset: int):
        """ print area add rows

        Args:
            st (_type_): sheet
            row_offset (int): rws

        Raises:
            ValueError: exception
        """
        if not st.print_area:
            return  # No print area

        if isinstance(st.print_area, (list, tuple)):  # list. eg. ['$B$2:$S$35']
            print_area_str = st.print_area[0]
        elif isinstance(st.print_area, str):  # str. eg. "'output'!$B$2:$S$35"
            print_area_str = st.print_area
        else:
            raise TypeError(f"Unsupported print_area type: {type(st.print_area)}")

        # get $B$2:$S$35 or B2:S35 format
        match = re.search(r"([$A-Z]+\$?\d+:[$A-Z]+\$?\d+)", print_area_str)
        if not match:
            raise ValueError(f"Invalid print_area format: {print_area_str}")

        coord = match.group(1).replace('$', '')
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row + row_offset}"

        sheet_match = re.match(r"(.*?)!?\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+", print_area_str)
        if sheet_match and "!" in sheet_match.group(0):
            sheet_name = sheet_match.group(1)
            new_range = f"{sheet_name}!{new_range}"

        st.print_area = new_range
